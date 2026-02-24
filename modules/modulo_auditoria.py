# modulo_auditoria.py
import os
import re
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.utils import column_index_from_string
import fitz  # PyMuPDF

def formatear_moneda_pdf(valor):
    if valor is None: 
        return None
    try:
        val_float = float(valor)
        if val_float == 0: 
            return None
        return "{:,.2f}".format(abs(val_float))
    except (ValueError, TypeError):
        return None

def indexar_pdfs_profundo(rutas):
    indice = defaultdict(list)
    patron_monto = re.compile(r'\d{1,3}(?:,\d{3})*\.\d{2}')
    for ruta in rutas:
        try:
            doc = fitz.open(ruta)
            for i, page in enumerate(doc):
                texto = page.get_text()
                matches = patron_monto.findall(texto)
                for monto in set(matches):
                    instancias = page.search_for(monto)
                    for rect in instancias:
                        if rect.x0 > 50: # Ajustado para capturar montos en más áreas
                            indice[monto].append({
                                "ruta": ruta, "pag": i, "rect": rect, "usado": False
                            })
            doc.close()
        except Exception as e:
            print(f"Error leyendo {os.path.basename(ruta)}: {e}")
    return indice

def ejecutar_auditoria(ruta_excel, dir_pdfs, dir_entregables):
    """
    Programa: Conciliacion IVA
    1. Busca IVA de AUX en Columna IVA de CFDI.
    2. Pega info en AUX.
    3. Busca TOTAL en PDFs.
    """
    try:
        lista_pdfs = [os.path.join(root, arc) 
                      for root, _, files in os.walk(dir_pdfs) 
                      for arc in files if arc.lower().endswith('.pdf')]
                      
        if not lista_pdfs:
            return False, "La carpeta de PDFs no existe o está vacía."

        wb = openpyxl.load_workbook(ruta_excel)
        
        # Intentar obtener hojas por nombre o por índice
        def get_sheet(wb, names):
            for name in names:
                if name in wb.sheetnames: return wb[name]
            return wb.worksheets[0]

        ws_aux = get_sheet(wb, ["AUX", "AUXILIAR", "AUX 2024"])
        ws_cfdi = get_sheet(wb, ["CFDI", "CFDI REC PROV"])

        # Identificar columnas en CFDI
        headers_cfdi = [str(cell.value).upper() if cell.value else "" for cell in ws_cfdi[5]] # Asumiendo fila 5 para CFDI
        if not any(headers_cfdi): headers_cfdi = [str(cell.value).upper() if cell.value else "" for cell in ws_cfdi[1]]

        idx_uuid = next((i for i, h in enumerate(headers_cfdi) if 'UUID' in h), 0)
        idx_iva_cfdi = next((i for i, h in enumerate(headers_cfdi) if 'IVA' in h), 1)
        idx_total_cfdi = next((i for i, h in enumerate(headers_cfdi) if 'TOTAL' in h), 2)

        # Identificar columnas en AUX
        headers_aux = [str(cell.value).upper() if cell.value else "" for cell in ws_aux[1]]
        idx_iva_aux = next((i for i, h in enumerate(headers_aux) if 'IVA' in h), 7) # Por defecto H(7)
        idx_total_aux_target = next((i for i, h in enumerate(headers_aux) if 'TOTAL' in h or 'MONTO' in h), 8) # Donde pegaremos el total

        # 1. Pre-cargar CFDI por Monto de IVA
        dict_iva_cfdi = {}
        for r_idx, row in enumerate(ws_cfdi.iter_rows(min_row=2, values_only=False), start=2):
            iva_val = formatear_moneda_pdf(row[idx_iva_cfdi].value)
            if iva_val:
                dict_iva_cfdi[iva_val] = row

        # 2. Indexar PDFs
        db_montos = indexar_pdfs_profundo(lista_pdfs)
        
        acciones_por_pdf = defaultdict(list)
        faltantes_reporte = [] 
        contador_ref = 1
        
        # 3. Recorrido AUX para Match de IVA y búsqueda de TOTAL en PDF
        for row_idx, row in enumerate(ws_aux.iter_rows(min_row=2, values_only=False), start=2):
            iva_aux = formatear_moneda_pdf(row[idx_iva_aux].value)
            
            if iva_aux and iva_aux in dict_iva_cfdi:
                row_c = dict_iva_cfdi[iva_aux]
                # Pegar el TOTAL del CFDI en el AUX (Paso solicitado)
                total_fiscal = row_c[idx_total_cfdi].value
                row[idx_total_aux_target].value = total_fiscal
                
                # Buscar ese TOTAL en los PDFs
                total_str = formatear_moneda_pdf(total_fiscal)
                if total_str and total_str in db_montos:
                    match_encontrado = next((m for m in db_montos[total_str] if not m["usado"]), None)
                    if match_encontrado:
                        match_encontrado["usado"] = True
                        acciones_por_pdf[match_encontrado["ruta"]].append({
                            "pag": match_encontrado["pag"],
                            "rect": match_encontrado["rect"], 
                            "ref": contador_ref
                        })
                        # Guardar referencia en el AUX
                        ref_col = len(row) - 1 # Usar ultima columna disponible para la referencia
                        row[ref_col].value = f"Ref:{contador_ref:03d}"
                        contador_ref += 1
            
            # Registro de faltantes si no se encontró en PDF
            last_col_val = row[len(row)-1].value
            if not last_col_val or "Ref:" not in str(last_col_val):
                monto_rep = row[idx_iva_aux].value or 0
                faltantes_reporte.append(f"Fila {row_idx} | IVA: {monto_rep}")

        # 4. Generar PDFs marcados
        pdfs_generados = 0
        for ruta_pdf, lista_acciones in acciones_por_pdf.items():
            try:
                doc = fitz.open(ruta_pdf)
                for accion in lista_acciones:
                    page = doc[accion["pag"]]
                    rect = accion["rect"]
                    annot = page.add_underline_annot(rect)
                    annot.set_colors(stroke=(0, 0.5, 0)) 
                    annot.update()
                    pt = fitz.Point(rect.x1 + 2, rect.y1)
                    page.insert_text(pt, f"Ref:{accion['ref']:03d}", fontsize=6, color=(0,0.5,0))
                
                nombre_salida = os.path.basename(ruta_pdf).replace(".pdf", "_IVA_AUDITADO.pdf")
                doc.save(os.path.join(dir_entregables, nombre_salida))
                pdfs_generados += 1
                doc.close()
            except Exception as e:
                print(f"Error marcando {ruta_pdf}: {e}")

        # 5. Reporte y Guardado
        ruta_txt = os.path.join(dir_entregables, "REPORTE_CONCILIACION_IVA.txt")
        with open(ruta_txt, "w", encoding="utf-8") as f:
            f.write(f"REPORTE CONCILIACIÓN IVA - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*50 + "\n")
            f.write(f"TOTAL MATCHES IVA (Fiscal vs Contable): {len(acciones_por_pdf)}\n")
            f.write(f"TOTAL TOTALES ENCONTRADOS EN PDF: {contador_ref - 1}\n")
            f.write("="*50 + "\n")
                
        ruta_final = os.path.join(dir_entregables, "CONCILIACION_IVA_FINAL.xlsx")
        wb.save(ruta_final)

        return True, f"Proceso Conciliación IVA exitoso. {pdfs_generados} PDFs generados."

    except Exception as e:
        import traceback
        return False, f"Error: {traceback.format_exc()}"