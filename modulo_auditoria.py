# modulo_auditoria.py
import os
import re
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.utils import column_index_from_string
import fitz  # PyMuPDF

def formatear_moneda_pdf(valor):
    if valor is None: 
        return None
    try:
        val_float = float(valor)
        if val_float == 0: 
            return None
        return "{:,.2f}".format(abs(val_float))
    except (ValueError, TypeError):
        return None

def indexar_pdfs_profundo(rutas):
    indice = defaultdict(list)
    patron_monto = re.compile(r'\d{1,3}(?:,\d{3})*\.\d{2}')
    for ruta in rutas:
        try:
            doc = fitz.open(ruta)
            for i, page in enumerate(doc):
                texto = page.get_text()
                matches = patron_monto.findall(texto)
                for monto in set(matches):
                    instancias = page.search_for(monto)
                    for rect in instancias:
                        if rect.x0 > 250: 
                            indice[monto].append({
                                "ruta": ruta, "pag": i, "rect": rect, "usado": False
                            })
            doc.close()
        except Exception as e:
            print(f"Error leyendo {os.path.basename(ruta)}: {e}")
    return indice

def ejecutar_auditoria(ruta_excel, dir_pdfs, dir_entregables):
    try:
        # 1. Búsqueda de PDFs optimizada
        lista_pdfs = [os.path.join(root, arc) 
                      for root, _, files in os.walk(dir_pdfs) 
                      for arc in files if arc.lower().endswith('.pdf')]
                      
        if not lista_pdfs:
            return False, "La carpeta de PDFs no existe o está vacía."

        # 2. Cargar Excel
        wb = openpyxl.load_workbook(ruta_excel)
        ws_aux = wb["AUX 2024"]
        ws_cfdi = wb["CFDI"]

        idx_CW = column_index_from_string('CW') - 1 
        idx_CX = column_index_from_string('CX') - 1
        
        db_montos = indexar_pdfs_profundo(lista_pdfs)
        
        filas_cfdi_borrar = set()
        acciones_por_pdf = defaultdict(list)
        faltantes_reporte = [] 
        contador_ref = 1

        # 3. Pre-cargar CFDI (Optimizado con Diccionario para máxima velocidad)
        diccionario_cfdi = {}
        for r_idx, row in enumerate(ws_cfdi.iter_rows(min_row=2, values_only=False), start=2):
            val82 = str(row[82].value).strip() if len(row) > 82 and row[82].value else ""
            val85 = str(row[85].value).strip() if len(row) > 85 and row[85].value else ""
            
            if val82: diccionario_cfdi[val82] = (r_idx, row)
            if val85: diccionario_cfdi[val85] = (r_idx, row)

        # 4. Recorrido AUX
        for row_idx, row in enumerate(ws_aux.iter_rows(min_row=4, values_only=False), start=4):
            if not row[0].value and not row[1].value: 
                continue

            celda_h = row[7]
            val_busqueda = str(celda_h.value).strip() if celda_h.value else None
            
            # Cruce instantáneo gracias al diccionario
            if val_busqueda and val_busqueda in diccionario_cfdi:
                f_idx, row_c = diccionario_cfdi[val_busqueda]
                if f_idx not in filas_cfdi_borrar:
                    for i, cell in enumerate(row_c):
                        row[10 + i].value = cell.value
                    filas_cfdi_borrar.add(f_idx)
            
            # Búsqueda de montos en PDF
            val_cw = row[idx_CW].value
            val_cx = row[idx_CX].value
            monto_str = formatear_moneda_pdf(val_cw) or formatear_moneda_pdf(val_cx)
                
            if monto_str and monto_str in db_montos:
                match_encontrado = next((m for m in db_montos[monto_str] if not m["usado"]), None)
                
                if match_encontrado:
                    match_encontrado["usado"] = True
                    acciones_por_pdf[match_encontrado["ruta"]].append({
                        "pag": match_encontrado["pag"],
                        "rect": match_encontrado["rect"], 
                        "ref": contador_ref
                    })
                    row[6].value = f"Ref:{contador_ref:03d}"
                    contador_ref += 1
            
            # Registro de faltantes
            if not row[6].value:
                desc = str(row[3].value)[:30] if row[3].value else "Sin Desc"
                monto_rep = val_cw or val_cx or 0
                faltantes_reporte.append(f"Fila {row_idx}: {desc} | Monto: {monto_rep}")

        # 5. Marcar PDFs
        pdfs_generados = 0
        for ruta_pdf, lista_acciones in acciones_por_pdf.items():
            if not lista_acciones: continue
            try:
                doc = fitz.open(ruta_pdf)
                for accion in lista_acciones:
                    page = doc[accion["pag"]]
                    rect = accion["rect"]
                    
                    # Subrayado verde
                    annot = page.add_underline_annot(rect)
                    annot.set_colors(stroke=(0, 0.5, 0)) 
                    annot.update()
                    
                    # Texto de referencia
                    pt = fitz.Point(rect.x1 + 2, rect.y1)
                    page.insert_text(pt, f"Ref:{accion['ref']:03d}", fontsize=5, color=(0,0.5,0))
                
                nombre_salida = os.path.basename(ruta_pdf).replace(".pdf", "_AUDITADO.pdf")
                doc.save(os.path.join(dir_entregables, nombre_salida))
                pdfs_generados += 1
                doc.close()
            except Exception as e:
                print(f"Error guardando PDF {ruta_pdf}: {e}")

        # 6. Limpieza CFDI
        data_backup = [row for row in ws_cfdi.iter_rows(values_only=True)]
        ws_cfdi.delete_rows(1, ws_cfdi.max_row)
        
        for i, row_data in enumerate(data_backup, start=1):
            if i not in filas_cfdi_borrar:
                ws_cfdi.append(row_data)

        # 7. Reporte Faltantes
        ruta_txt = os.path.join(dir_entregables, "REPORTE_DETALLADO_FALTANTES.txt")
        with open(ruta_txt, "w", encoding="utf-8") as f:
            f.write(f"REPORTE DE AUDITORÍA - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*50 + "\n")
            f.write(f"TOTAL MOVIMIENTOS ENCONTRADOS EN PDF: {contador_ref - 1}\n")
            f.write(f"TOTAL MOVIMIENTOS RESTANTES EN AUXILIAR: {len(faltantes_reporte)}\n")
            f.write("="*50 + "\n\n")
            f.write("LISTADO DE MOVIMIENTOS DEL AUXILIAR SIN MATCH EN PDF:\n")
            for item in faltantes_reporte:
                f.write(item + "\n")
                
        # 8. Guardar Excel Final
        ruta_final = os.path.join(dir_entregables, "GSM_FINAL_V4.xlsx")
        wb.save(ruta_final)

        return True, f"Proceso exitoso: {pdfs_generados} PDFs marcados y {len(faltantes_reporte)} faltantes reportados."

    except Exception as e:
        import traceback
        return False, f"Error en módulo auditoría: {traceback.format_exc()}"